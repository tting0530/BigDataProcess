#!/bin/user/python2

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb["Sheet1"]

row_id = 1
#cnt = 0
#for r in ws:
#	cnt += 1
#row_id = 1
for row in ws:
	if row_id != 1:
		sum = ws.cell(row = row_id, column = 3).value * 0.3
		sum += ws.cell(row = row_id, column = 4).value * 0.35
		sum += ws.cell(row = row_id, column = 5).value * 0.34
		sum += ws.cell(row = row_id, column = 6).value * 0.3
		ws.cell(row = row_id, column = 7).value = sum
	row_id += 1

row_id = 1
#grade
for row in ws:
	if row_id != 1:
		if row_id < cnt/3:
			ws.cell(row = row_id, column = 8).value = 'A'
		
		elif row_id < cnt/5:
			ws.cell(row = row_id, column = 8).value = 'B'
		else:
			ws.cell(row = row_id, column = 8).value = 'C'

wb.save("student.xlsx")
