#!/bin/user/python2

import openpyxl
wb = openpyxl.load_workbook("student.xlsx")
ws = wb["Sheet1"]


row_id = 1
cnt = 0
dic = {}

for r in ws:
	cnt += 1
row_id = 1
for row in ws:
	if row_id != 1:
		sum = ws.cell(row = row_id, column = 3).value * 0.3
		sum += ws.cell(row = row_id, column = 4).value * 0.35
		sum += ws.cell(row = row_id, column = 5).value * 0.34
		sum += ws.cell(row = row_id, column = 6).value * 0.3
		ws.cell(row = row_id, column = 7).value = sum
		dic[row_id] = sum
	row_id += 1

sorted_dict = dict(sorted(dic.items(), key = lambda item: item[1]))
print(sorted_dict)


row_id = 1
print("cnt:", cnt)
#grade
aCnt=int(cnt*0.3)
bCnt=int(cnt*0.7)
cCnt=cnt-aCnt-bCnt
aPlusCnt=int(aCnt/2)
bPlusCnt=int(bCnt/2)
cPlusCnt=int(cCnt/2)
print("cnt/3", cnt/3)
row_id=1


for row_id in len(sorted_dict):
	if row_id < aCnt:
		if row_id < aPlusCnt:
			sorted_dict[row_id]="A+"
		else:
			sorted_dict[row_id]="A0"
	elif row_id < bCnt:
		if row_id < bPlusCnt:
			sorted_dict[row_id]="B+"
		else:
			sorted_dict[row_id]="B0"
	else:	
		if row_id < cPlusCnt:
			sorted_dict[row_id]="C+"
		else:
			sorted_dict[row_id]="C0"

print(dict)
row_id = 1
for row in ws:
	if row_id != 1:
		ws.cell(row = row_id, column = 9).value= 0
	row_id += 1

wb.save("student.xlsx")
