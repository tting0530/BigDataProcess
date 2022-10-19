#!/bin/user/python2

import openpyxl
wb = openpyxl.load_workbook("student.xlsx")
ws = wb["Sheet1"]

row_id = 1
len = 0
dic = {}

for r in ws:
	len += 1
len=len-1
row_id = 1
for row in ws:
	if row_id != 1:
		sum = ws.cell(row = row_id, column = 3).value * 0.3
		sum += ws.cell(row = row_id, column = 4).value * 0.35
		sum += ws.cell(row = row_id, column = 5).value * 0.34
		sum += ws.cell(row = row_id, column = 6).value * 0.3
		ws.cell(row = row_id, column = 7).value = sum
		dic[row_id] = sum
	row_id += 1

print("original dic\n", dic, "\n")

sorted_dict = dict(sorted(dic.items(),reverse=True,  key = lambda item: item[1]))
print(sorted_dict)


row_id = 1
#grade
aCnt=int(len*0.3)
print(len , "bCnt = len*0.3 - aCnt ", "bCnt", "=", len*0.7, " - ", aCnt)
bCnt=int(len*0.7-aCnt)
print("bCnt = len*0.3 - aCnt ", bCnt, "=", len*0.7, " - ", aCnt)
cCnt=len-aCnt-bCnt
print("len:" , len)
print("aCnt: " , aCnt)
print("bCnt: " , bCnt)
print("cCnt: " , cCnt)
aPlusCnt=int(aCnt/2)
bPlusCnt=int(bCnt/2)
cPlusCnt=int(cCnt/2)
row_id=1
cnt=0
j=0
for row_id in sorted_dict.keys():
#	print("now grade check row_id(=key):", row_id, "\n")
	if cnt < aCnt:
		if cnt < aPlusCnt:
			sorted_dict[row_id]="A+"
		else:
			sorted_dict[row_id]="A0"
	elif cnt < bCnt+aCnt:
		if cnt < bPlusCnt+aCnt:
			sorted_dict[row_id]="B+"
		else:
			sorted_dict[row_id]="B0"
	else:
		print("cnt:",cnt , " cPlusCnt" , cPlusCnt)	
		if cnt < cPlusCnt+aCnt+bCnt:
			sorted_dict[row_id]="C+"
		else:
			sorted_dict[row_id]="C0"
	cnt += 1

#	ws.cell(row = row_id, column = 8).value=sorted_dict[row_id] 
#	print(ws.cell(row = row_id, column = 8).value) 
#print(sorted_dict)

for i in sorted_dict.keys():
	j+=1
	print(i, "value:", sorted_dict[i], "순위:",j)
	ws.cell(row = i, column = 8).value=sorted_dict[i]

print(ws.cell(row = 46, column = 8).value)
wb.save("student.xlsx")
