#!/bin/user/python2

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb["Sheet1"]

row_id = 1;
for row in ws:
	if row_id != 1:
		sum = ws.cell(row = row_id, column = 3).value * 0.3
		sum += ws.cell(row = row_id, column = 4).value * 0.35
		sum += ws.cell(row = row_id, column = 5).value * 0.34
		sum += ws.cell(row = row_id, column = 6).value * 0.3
		ws.cell(row = row_id, column = 7).value = sum
	row_id += 1
#grade

wb.save("student.xlsx")
